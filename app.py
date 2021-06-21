
import streamlit as st
from openpyxl import load_workbook
@st.cache

def about():
    st.write(
        '''
        contact email - tanbinhasnat04@gmail.com
        ''')


def main():
	st.set_page_config(layout="wide")
	
	def caln():
		prop_arr=[]
		for i in range(1,86):
			name_tr=sheet.cell(row=1,column=i).value
			prop_arr.append(name_tr)
		if section in name_tr_label:
			indi_name=name_tr_label.index(section)+1
		j=0
		for i in range(4,89):
			st.write(f'{prop_arr[j]} = {sheet.cell(row=indi_name,column=i-4).value}')
			j+=1
		
		
		
	

	st.title("Aisc database version - v15 : ")
    

	activities = ["Home", "About"]
	choice = st.sidebar.selectbox("Pick something fun", activities)

	if choice == "Home":

        
		wb=load_workbook('aisc.xlsx')
		sheet=wb.active
		names=[]

	
		name_tr_label=[]
		for i in range(1,len(sheet['A'])+1):
			if i==1:
				name_tr_label.append('')
			else:
				name_tr=sheet.cell(row=i,column=3).value
				name_tr_label.append(name_tr)
		section=st.selectbox('Select or write section',name_tr_label)
		caln()

		
        

	elif choice == "About":
		
		about()



if __name__ == "__main__":
	main()

